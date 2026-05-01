#!/usr/bin/env python3
from __future__ import annotations

import argparse
import copy
import json
import os
import re
import shutil
import sys
import time
import uuid
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import requests
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


REPO_ROOT = Path(os.environ.get("HERMES_DEMO_REPO_ROOT", Path(__file__).resolve().parents[1])).resolve()
DEMO_DIR = REPO_ROOT / "output" / "spreadsheet" / "hermes-demo-pack"
WORKBOOK_PATH = DEMO_DIR / "hermes_demo_pack.xlsx"
RESULTS_DIR = DEMO_DIR / "test-results"
STATEFUL_WORKBOOK_PATH = RESULTS_DIR / "hermes_demo_pack.smoke.xlsx"

GATEWAY_BASE_URL = os.environ.get("HERMES_DEMO_GATEWAY_URL", "http://127.0.0.1:8789").rstrip("/")
HOST_PLATFORM = os.environ.get("HERMES_DEMO_HOST_PLATFORM", "google_sheets")
SOURCE_CHANNEL = os.environ.get("HERMES_DEMO_SOURCE_CHANNEL", HOST_PLATFORM)
CLIENT_VERSION = os.environ.get("HERMES_DEMO_CLIENT_VERSION", "demo-smoke-harness")
WORKBOOK_SESSION_KEY = os.environ.get(
    "HERMES_DEMO_WORKBOOK_SESSION_KEY",
    f"{HOST_PLATFORM}::hermes-demo-pack"
)
POLL_INTERVAL_SECONDS = float(os.environ.get("HERMES_DEMO_POLL_INTERVAL_SECONDS", "1.0"))
POLL_TIMEOUT_SECONDS = float(os.environ.get("HERMES_DEMO_POLL_TIMEOUT_SECONDS", "180.0"))

MAX_CONTEXT_CELLS = 400
MAX_SHEET_PREVIEW_ROWS = 5
MAX_SHEET_PREVIEW_SHEETS = 8
MAX_REFERENCED_CELLS = 20
MAX_CONTEXT_TEXT_LENGTH = 4000
MAX_FORMULA_LENGTH = 16000

A1_RE = re.compile(r"\$?([A-Z]{1,3})\$?([1-9][0-9]*)")


@dataclass(frozen=True)
class DemoCase:
    id: str
    sheet: str
    prompt: str
    selection: str
    active_cell: str
    image_path: Path | None = None
    expect_write: bool = False
    preferred_types: tuple[str, ...] = ()
    notes: str = ""


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def truncate_text(value: Any, max_length: int = MAX_CONTEXT_TEXT_LENGTH) -> Any:
    if not isinstance(value, str):
        return value
    if len(value) <= max_length:
        return value
    return f"{value[: max_length - 1]}…"


def column_letters_to_index(column_letters: str) -> int:
    total = 0
    for character in column_letters.strip().upper():
        total = (total * 26) + (ord(character) - 64)
    return total


def column_index_to_letters(index: int) -> str:
    value = int(index)
    letters = ""
    while value > 0:
        value, remainder = divmod(value - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def split_sheet_and_range(a1: str) -> tuple[str | None, str]:
    if "!" not in a1:
        return None, a1
    sheet_name, range_ref = a1.split("!", 1)
    return sheet_name.strip("'"), range_ref


def parse_a1_range(a1: str) -> dict[str, int]:
    _, range_ref = split_sheet_and_range(a1)
    normalized = range_ref.replace("$", "").strip()
    if ":" in normalized:
        start_ref, end_ref = normalized.split(":", 1)
    else:
        start_ref = end_ref = normalized
    start_match = A1_RE.fullmatch(start_ref)
    end_match = A1_RE.fullmatch(end_ref)
    if not start_match or not end_match:
        raise ValueError(f"Unsupported A1 reference: {a1}")
    start_col, start_row = start_match.groups()
    end_col, end_row = end_match.groups()
    start_column = column_letters_to_index(start_col)
    end_column = column_letters_to_index(end_col)
    start_row_num = int(start_row)
    end_row_num = int(end_row)
    if start_column > end_column:
        start_column, end_column = end_column, start_column
    if start_row_num > end_row_num:
        start_row_num, end_row_num = end_row_num, start_row_num
    return {
        "start_row": start_row_num,
        "end_row": end_row_num,
        "start_column": start_column,
        "end_column": end_column,
        "row_count": end_row_num - start_row_num + 1,
        "column_count": end_column - start_column + 1,
    }


def build_a1_range(
    start_row: int,
    start_column: int,
    end_row: int,
    end_column: int,
) -> str:
    start_ref = f"{column_index_to_letters(start_column)}{start_row}"
    end_ref = f"{column_index_to_letters(end_column)}{end_row}"
    return start_ref if start_ref == end_ref else f"{start_ref}:{end_ref}"


def shape_from_range(a1: str) -> dict[str, int]:
    bounds = parse_a1_range(a1)
    return {
        "rows": bounds["row_count"],
        "columns": bounds["column_count"],
    }


def normalize_formula(value: Any) -> str | None:
    if isinstance(value, str) and value.startswith("="):
        return truncate_text(value, MAX_FORMULA_LENGTH)
    return None


def normalize_cell_value(cell: Cell) -> Any:
    value = cell.value
    if isinstance(value, str):
        return truncate_text(value)
    return value


def extract_headers(matrix: list[list[Any]]) -> list[str] | None:
    if not matrix:
        return None
    first_row = matrix[0]
    if not first_row:
        return None
    headers: list[str] = []
    for cell in first_row:
        if not isinstance(cell, str) or not cell.strip():
            return None
        headers.append(truncate_text(cell, 256))
    return headers or None


def find_used_bounds(sheet: Worksheet) -> dict[str, int]:
    min_row: int | None = None
    min_col: int | None = None
    max_row = 1
    max_col = 1
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            min_row = cell.row if min_row is None else min(min_row, cell.row)
            min_col = cell.column if min_col is None else min(min_col, cell.column)
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
    if min_row is None or min_col is None:
        return {
            "start_row": 1,
            "end_row": 1,
            "start_column": 1,
            "end_column": 1,
            "row_count": 1,
            "column_count": 1,
        }
    return {
        "start_row": min_row,
        "end_row": max_row,
        "start_column": min_col,
        "end_column": max_col,
        "row_count": max_row - min_row + 1,
        "column_count": max_col - min_col + 1,
    }


def build_matrix_from_bounds(sheet: Worksheet, bounds: dict[str, int]) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(
        min_row=bounds["start_row"],
        max_row=bounds["end_row"],
        min_col=bounds["start_column"],
        max_col=bounds["end_column"],
    ):
        rows.append([normalize_cell_value(cell) for cell in row])
    return rows


def build_formula_matrix_from_bounds(sheet: Worksheet, bounds: dict[str, int]) -> list[list[str | None]]:
    rows: list[list[str | None]] = []
    for row in sheet.iter_rows(
        min_row=bounds["start_row"],
        max_row=bounds["end_row"],
        min_col=bounds["start_column"],
        max_col=bounds["end_column"],
    ):
        rows.append([normalize_formula(cell.value) for cell in row])
    return rows


def should_include_matrix(bounds: dict[str, int]) -> bool:
    return bounds["row_count"] * bounds["column_count"] <= MAX_CONTEXT_CELLS


def get_cell_context(sheet: Worksheet, a1: str) -> dict[str, Any]:
    cell = sheet[a1]
    context: dict[str, Any] = {
        "a1Notation": a1,
        "displayValue": truncate_text("" if cell.value is None else str(cell.value)),
        "value": normalize_cell_value(cell),
    }
    formula = normalize_formula(cell.value)
    if formula:
        context["formula"] = formula
    return context


def extract_referenced_a1(prompt: str) -> list[str]:
    found: list[str] = []
    for match in re.finditer(r"\b([A-Z]{1,3}\d{1,7})\b", prompt.upper()):
        ref = match.group(1)
        if ref not in found:
            found.append(ref)
    return found[:MAX_REFERENCED_CELLS]


class WorkbookModel:
    def __init__(self, workbook_path: Path):
        self.path = workbook_path
        self.workbook = openpyxl.load_workbook(workbook_path, data_only=False)

    def save(self) -> None:
        self.workbook.save(self.path)

    def sheet(self, name: str) -> Worksheet:
        return self.workbook[name]

    def ensure_sheet(self, name: str) -> Worksheet:
        if name in self.workbook.sheetnames:
            return self.workbook[name]
        return self.workbook.create_sheet(name)

    def used_range(self, sheet_name: str) -> str:
        bounds = find_used_bounds(self.sheet(sheet_name))
        return build_a1_range(
            bounds["start_row"],
            bounds["start_column"],
            bounds["end_row"],
            bounds["end_column"],
        )

    def build_request_context(self, case: DemoCase) -> dict[str, Any]:
        sheet = self.sheet(case.sheet)
        selection_bounds = parse_a1_range(case.selection)
        current_region_a1 = self.used_range(case.sheet)
        current_region_bounds = parse_a1_range(current_region_a1)

        selection_context: dict[str, Any] = {"range": case.selection}
        if should_include_matrix(selection_bounds):
            selection_values = build_matrix_from_bounds(sheet, selection_bounds)
            selection_formulas = build_formula_matrix_from_bounds(sheet, selection_bounds)
            selection_context["values"] = selection_values
            selection_context["formulas"] = selection_formulas
            headers = extract_headers(selection_values)
            if headers:
                selection_context["headers"] = headers
        else:
            header_bounds = {
                **selection_bounds,
                "end_row": selection_bounds["start_row"],
                "row_count": 1,
            }
            headers = extract_headers(build_matrix_from_bounds(sheet, header_bounds))
            if headers:
                selection_context["headers"] = headers

        current_region_context: dict[str, Any] = {"range": current_region_a1}
        if should_include_matrix(current_region_bounds):
            region_values = build_matrix_from_bounds(sheet, current_region_bounds)
            region_formulas = build_formula_matrix_from_bounds(sheet, current_region_bounds)
            current_region_context["values"] = region_values
            current_region_context["formulas"] = region_formulas
            headers = extract_headers(region_values)
            if headers:
                current_region_context["headers"] = headers
        else:
            header_bounds = {
                **current_region_bounds,
                "end_row": current_region_bounds["start_row"],
                "row_count": 1,
            }
            headers = extract_headers(build_matrix_from_bounds(sheet, header_bounds))
            if headers:
                current_region_context["headers"] = headers

        active_cell_context = get_cell_context(sheet, case.active_cell)
        referenced_cells = [
            get_cell_context(sheet, ref)
            for ref in extract_referenced_a1(case.prompt)
            if ref != case.active_cell and ref in sheet
        ]
        implicit_targets = {
            "currentRegionArtifactTarget": build_a1_range(
                current_region_bounds["end_row"] + 2,
                current_region_bounds["start_column"],
                current_region_bounds["end_row"] + 2,
                current_region_bounds["start_column"],
            ),
            "currentRegionAppendTarget": build_a1_range(
                current_region_bounds["end_row"] + 1,
                current_region_bounds["start_column"],
                current_region_bounds["end_row"] + 1,
                current_region_bounds["end_column"],
            ),
        }

        context: dict[str, Any] = {
            "selection": selection_context,
            "currentRegion": current_region_context,
            "activeCell": active_cell_context,
            **implicit_targets,
            "sheetsPreview": self.build_sheets_preview(),
        }
        if referenced_cells:
            context["referencedCells"] = referenced_cells
        return context

    def build_sheets_preview(self) -> list[dict[str, Any]]:
        previews: list[dict[str, Any]] = []
        for sheet_name in self.workbook.sheetnames[:MAX_SHEET_PREVIEW_SHEETS]:
            sheet = self.sheet(sheet_name)
            bounds = find_used_bounds(sheet)
            preview_bounds = copy.deepcopy(bounds)
            preview_bounds["end_row"] = min(bounds["end_row"], bounds["start_row"] + MAX_SHEET_PREVIEW_ROWS - 1)
            preview_bounds["row_count"] = preview_bounds["end_row"] - preview_bounds["start_row"] + 1
            values = build_matrix_from_bounds(sheet, preview_bounds)
            preview: dict[str, Any] = {"sheetName": sheet_name}
            headers = extract_headers(values)
            if headers:
                preview["headers"] = headers
            if values:
                preview["values"] = values
            previews.append(preview)
        return previews

    def _target_cells(self, sheet_name: str, target_range: str) -> tuple[Worksheet, dict[str, int]]:
        sheet = self.ensure_sheet(sheet_name)
        bounds = parse_a1_range(target_range)
        return sheet, bounds

    def _write_matrix(
        self,
        sheet_name: str,
        target_range: str,
        *,
        values: list[list[Any]] | None = None,
        formulas: list[list[str | None]] | None = None,
    ) -> None:
        sheet, bounds = self._target_cells(sheet_name, target_range)
        for row_offset in range(bounds["row_count"]):
            for col_offset in range(bounds["column_count"]):
                cell = sheet.cell(bounds["start_row"] + row_offset, bounds["start_column"] + col_offset)
                formula_value = formulas[row_offset][col_offset] if formulas else None
                if isinstance(formula_value, str) and formula_value.strip():
                    cell.value = formula_value
                    continue
                raw_value = values[row_offset][col_offset] if values else None
                cell.value = raw_value

    def _append_rows(self, sheet_name: str, target_range: str, values: list[list[Any]]) -> str:
        sheet, bounds = self._target_cells(sheet_name, target_range)
        first_empty_row = bounds["start_row"]
        for row_index in range(bounds["start_row"], bounds["end_row"] + 1):
            row_values = [
                sheet.cell(row_index, col_index).value
                for col_index in range(bounds["start_column"], bounds["end_column"] + 1)
            ]
            if all(value in (None, "") for value in row_values):
                first_empty_row = row_index
                break
            first_empty_row = row_index + 1
        end_row = first_empty_row + len(values) - 1
        actual_range = build_a1_range(
            first_empty_row,
            bounds["start_column"],
            end_row,
            bounds["start_column"] + len(values[0]) - 1,
        )
        self._write_matrix(sheet_name, actual_range, values=values)
        return actual_range

    def _cleanup_transform(self, plan: dict[str, Any]) -> None:
        sheet, bounds = self._target_cells(plan["targetSheet"], plan["targetRange"])
        matrix = build_matrix_from_bounds(sheet, bounds)
        operation = plan["operation"]

        if operation == "trim_whitespace":
            matrix = [
                [value.strip() if isinstance(value, str) else value for value in row]
                for row in matrix
            ]
        elif operation == "remove_blank_rows":
            matrix = [
                row for row in matrix
                if any(value not in (None, "") for value in row)
            ]
        elif operation == "remove_duplicate_rows":
            headers = extract_headers(matrix) or []
            key_columns = plan.get("keyColumns") or headers or [
                column_index_to_letters(index + 1) for index in range(len(matrix[0]) if matrix else 0)
            ]
            header_index = {header: idx for idx, header in enumerate(headers)}
            key_indices: list[int] = []
            for key_column in key_columns:
                if isinstance(key_column, str) and key_column in header_index:
                    key_indices.append(header_index[key_column])
                elif isinstance(key_column, str) and key_column.isalpha():
                    key_indices.append(column_letters_to_index(key_column) - 1)
            seen: set[tuple[Any, ...]] = set()
            deduped: list[list[Any]] = []
            for row_index, row in enumerate(matrix):
                if row_index == 0 and headers:
                    deduped.append(row)
                    continue
                key = tuple(row[index] if index < len(row) else None for index in key_indices) if key_indices else tuple(row)
                if key in seen:
                    continue
                seen.add(key)
                deduped.append(row)
            matrix = deduped
        elif operation == "normalize_case":
            mode = plan.get("mode", "title")

            def normalize_case(value: Any) -> Any:
                if not isinstance(value, str):
                    return value
                if mode == "upper":
                    return value.upper()
                if mode == "lower":
                    return value.lower()
                return value.title()

            matrix = [[normalize_case(value) for value in row] for row in matrix]
        elif operation == "fill_down":
            columns = plan.get("columns") or []
            headers = extract_headers(matrix) or []
            target_indices = set()
            for column_name in columns:
                if column_name in headers:
                    target_indices.add(headers.index(column_name))
                elif isinstance(column_name, str) and column_name.isalpha():
                    target_indices.add(column_letters_to_index(column_name) - 1)
            if not target_indices and matrix:
                target_indices = set(range(len(matrix[0])))
            for column_index in target_indices:
                last_value: Any = None
                for row_index, row in enumerate(matrix):
                    if row_index == 0 and headers:
                        continue
                    if column_index >= len(row):
                        continue
                    value = row[column_index]
                    if value in (None, ""):
                        row[column_index] = last_value
                    else:
                        last_value = value
        elif operation == "standardize_format":
            format_type = plan.get("formatType")
            if format_type == "number_text":
                matrix = [
                    [
                        float(str(value).replace(",", "").replace("$", "")) if isinstance(value, str) and str(value).strip() else value
                        for value in row
                    ]
                    for row in matrix
                ]
        elif operation == "join_columns":
            delimiter = plan.get("delimiter", " ")
            source_columns = plan.get("sourceColumns") or []
            target_column = plan.get("targetColumn")
            headers = extract_headers(matrix) or []
            source_indices = []
            for source_column in source_columns:
                if source_column in headers:
                    source_indices.append(headers.index(source_column))
                elif source_column.isalpha():
                    source_indices.append(column_letters_to_index(source_column) - 1)
            if target_column in headers:
                target_index = headers.index(target_column)
            else:
                target_index = column_letters_to_index(str(target_column)) - 1
            for row_index, row in enumerate(matrix):
                if row_index == 0 and headers:
                    continue
                joined = delimiter.join(str(row[index]) for index in source_indices if index < len(row) and row[index] not in (None, ""))
                while len(row) <= target_index:
                    row.append("")
                row[target_index] = joined
        elif operation == "split_column":
            delimiter = plan.get("delimiter", " ")
            source_column = plan.get("sourceColumn")
            target_start_column = plan.get("targetStartColumn")
            headers = extract_headers(matrix) or []
            if source_column in headers:
                source_index = headers.index(source_column)
            else:
                source_index = column_letters_to_index(str(source_column)) - 1
            target_index = column_letters_to_index(str(target_start_column)) - 1
            for row_index, row in enumerate(matrix):
                if source_index >= len(row):
                    continue
                parts = str(row[source_index]).split(delimiter)
                while len(row) < target_index + len(parts):
                    row.append("")
                for offset, part in enumerate(parts):
                    row[target_index + offset] = part

        target_range = build_a1_range(
            bounds["start_row"],
            bounds["start_column"],
            bounds["start_row"] + max(len(matrix), 1) - 1,
            bounds["start_column"] + max(len(matrix[0]) if matrix else bounds["column_count"], 1) - 1,
        )
        self._write_matrix(plan["targetSheet"], target_range, values=matrix or [[""]])

    def apply_plan(self, response_type: str, plan: dict[str, Any], execution_id: str | None = None) -> dict[str, Any]:
        host_platform = HOST_PLATFORM
        if response_type == "composite_plan":
            step_results: list[dict[str, Any]] = []
            for step in plan["steps"]:
                step_type = infer_plan_type(step["plan"])
                try:
                    self.apply_plan(step_type, step["plan"], execution_id=execution_id)
                    step_results.append({
                        "stepId": step["stepId"],
                        "status": "completed",
                        "summary": summarize_step(step_type, step["plan"]),
                    })
                except Exception as exc:  # pragma: no cover - defensive harness
                    step_results.append({
                        "stepId": step["stepId"],
                        "status": "failed",
                        "summary": str(exc),
                    })
                    if not step.get("continueOnError"):
                        raise
            return {
                "kind": "composite_update",
                "hostPlatform": host_platform,
                "operation": "composite_update",
                "executionId": execution_id or f"exec_{uuid.uuid4().hex[:12]}",
                "stepResults": step_results,
                "summary": f"Completed {len(step_results)} workflow step(s).",
            }

        if response_type == "sheet_update":
            target_range = plan["targetRange"]
            if plan["operation"] == "append_rows":
                target_range = self._append_rows(plan["targetSheet"], plan["targetRange"], plan["values"])
            elif plan["operation"] == "set_formulas":
                self._write_matrix(plan["targetSheet"], plan["targetRange"], formulas=plan["formulas"])
            else:
                self._write_matrix(
                    plan["targetSheet"],
                    plan["targetRange"],
                    values=plan.get("values"),
                    formulas=plan.get("formulas"),
                )
            return {
                "kind": "range_write",
                "hostPlatform": host_platform,
                "targetSheet": plan["targetSheet"],
                "targetRange": target_range,
                "writtenRows": plan["shape"]["rows"],
                "writtenColumns": plan["shape"]["columns"],
            }

        if response_type == "sheet_import_plan":
            values = [plan["headers"], *plan["values"]]
            self._write_matrix(plan["targetSheet"], plan["targetRange"], values=values)
            return {
                "kind": "range_write",
                "hostPlatform": host_platform,
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "writtenRows": plan["shape"]["rows"],
                "writtenColumns": plan["shape"]["columns"],
            }

        if response_type == "range_format_update":
            shape = plan.get("shape") or shape_from_range(plan["targetRange"])
            return {
                "kind": "range_write",
                "hostPlatform": host_platform,
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "writtenRows": shape["rows"],
                "writtenColumns": shape["columns"],
            }

        if response_type == "workbook_structure_update":
            operation = plan["operation"]
            if operation == "create_sheet":
                self.ensure_sheet(plan["sheetName"])
            elif operation == "delete_sheet":
                if plan["sheetName"] in self.workbook.sheetnames and len(self.workbook.sheetnames) > 1:
                    sheet = self.workbook[plan["sheetName"]]
                    self.workbook.remove(sheet)
            elif operation == "rename_sheet":
                self.workbook[plan["sheetName"]].title = plan["newSheetName"]
            elif operation == "duplicate_sheet":
                copied = self.workbook.copy_worksheet(self.workbook[plan["sheetName"]])
                if plan.get("newSheetName"):
                    copied.title = plan["newSheetName"]
            return {
                "kind": "workbook_structure_update",
                "hostPlatform": host_platform,
                "sheetName": plan["sheetName"],
                "operation": operation,
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "sheet_structure_update":
            return {
                "kind": "sheet_structure_update",
                "hostPlatform": host_platform,
                "targetSheet": plan["targetSheet"],
                "operation": plan["operation"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "range_sort_plan":
            return {
                "kind": "range_sort",
                "hostPlatform": host_platform,
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "range_filter_plan":
            return {
                "kind": "range_filter",
                "hostPlatform": host_platform,
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "data_validation_plan":
            return {
                "kind": "data_validation_update",
                "hostPlatform": host_platform,
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "conditional_format_plan":
            return {
                "kind": "conditional_format_update",
                "hostPlatform": host_platform,
                "operation": "conditional_format_update",
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "managementMode": plan["managementMode"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "named_range_update":
            return {
                "kind": "named_range_update",
                "hostPlatform": host_platform,
                "operation": plan["operation"],
                "name": plan["name"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "range_transfer_plan":
            source_sheet, source_bounds = self._target_cells(plan["sourceSheet"], plan["sourceRange"])
            values = build_matrix_from_bounds(source_sheet, source_bounds)
            self._write_matrix(plan["targetSheet"], plan["targetRange"], values=values)
            return {
                "kind": "range_transfer_update",
                "hostPlatform": host_platform,
                "operation": "range_transfer_update",
                "sourceSheet": plan["sourceSheet"],
                "sourceRange": plan["sourceRange"],
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "transferOperation": plan["operation"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "data_cleanup_plan":
            self._cleanup_transform(plan)
            return {
                "kind": "data_cleanup_update",
                "hostPlatform": host_platform,
                "operation": "data_cleanup_update",
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "cleanupOperation": plan["operation"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "analysis_report_plan":
            target_sheet = self.ensure_sheet(plan["targetSheet"])
            target_sheet[split_sheet_and_range(plan["targetRange"])[1].split(":")[0]] = plan["sections"][0]["summary"]
            return {
                "kind": "analysis_report_update",
                "hostPlatform": host_platform,
                "operation": "analysis_report_update",
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "pivot_table_plan":
            self.ensure_sheet(plan["targetSheet"])
            return {
                "kind": "pivot_table_update",
                "hostPlatform": host_platform,
                "operation": "pivot_table_update",
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "summary": summarize_step(response_type, plan),
            }

        if response_type == "chart_plan":
            self.ensure_sheet(plan["targetSheet"])
            return {
                "kind": "chart_update",
                "hostPlatform": host_platform,
                "operation": "chart_update",
                "targetSheet": plan["targetSheet"],
                "targetRange": plan["targetRange"],
                "chartType": plan["chartType"],
                "summary": summarize_step(response_type, plan),
            }

        raise ValueError(f"Unsupported plan family for local workbook simulation: {response_type}")


def summarize_step(plan_type: str, plan: dict[str, Any]) -> str:
    if plan_type == "workbook_structure_update":
        return f"{plan['operation']} on sheet {plan['sheetName']}."
    if "targetSheet" in plan and "targetRange" in plan:
        return f"{plan_type} on {plan['targetSheet']}!{plan['targetRange']}."
    if "targetSheet" in plan:
        return f"{plan_type} on {plan['targetSheet']}."
    return f"Applied {plan_type}."


def infer_plan_type(plan: dict[str, Any]) -> str:
    if "steps" in plan:
        return "composite_plan"
    if "sourceAttachmentId" in plan and "headers" in plan and "values" in plan:
        return "sheet_import_plan"
    if "chartType" in plan and "series" in plan:
        return "chart_plan"
    if "rowGroups" in plan and "valueAggregations" in plan:
        return "pivot_table_plan"
    if "outputMode" in plan and "sections" in plan:
        return "analysis_report_plan"
    if "managementMode" in plan and "ruleType" in plan:
        return "conditional_format_plan"
    if "ruleType" in plan and "invalidDataBehavior" in plan:
        return "data_validation_plan"
    if "sourceSheet" in plan and "sourceRange" in plan and "pasteMode" in plan:
        return "range_transfer_plan"
    if plan.get("operation") in {"replace_range", "append_rows", "set_formulas", "set_notes", "mixed_update"}:
        return "sheet_update"
    if plan.get("operation") in {
        "create_sheet",
        "delete_sheet",
        "rename_sheet",
        "duplicate_sheet",
        "move_sheet",
        "hide_sheet",
        "unhide_sheet",
    }:
        return "workbook_structure_update"
    if plan.get("operation") in {
        "insert_rows",
        "delete_rows",
        "hide_rows",
        "unhide_rows",
        "group_rows",
        "ungroup_rows",
        "insert_columns",
        "delete_columns",
        "hide_columns",
        "unhide_columns",
        "group_columns",
        "ungroup_columns",
        "merge_cells",
        "unmerge_cells",
        "freeze_panes",
        "unfreeze_panes",
        "autofit_rows",
        "autofit_columns",
        "set_sheet_tab_color",
    }:
        return "sheet_structure_update"
    if "keys" in plan and "hasHeader" in plan:
        return "range_sort_plan"
    if "conditions" in plan and "combiner" in plan:
        return "range_filter_plan"
    if plan.get("operation") in {
        "trim_whitespace",
        "remove_blank_rows",
        "remove_duplicate_rows",
        "normalize_case",
        "split_column",
        "join_columns",
        "fill_down",
        "standardize_format",
    }:
        return "data_cleanup_plan"
    if "scope" in plan and "name" in plan:
        return "named_range_update"
    if "format" in plan and "targetSheet" in plan and "targetRange" in plan:
        return "range_format_update"
    raise ValueError(f"Cannot infer plan family from step payload keys: {sorted(plan.keys())}")


def expectation_issue(case: DemoCase, response_type: str) -> str | None:
    if case.preferred_types and response_type not in case.preferred_types:
        return f"Expected one of {', '.join(case.preferred_types)} but got {response_type}."
    if case.expect_write and response_type in {"chat", "formula"}:
        return f"Prompt looks write-intent but Hermes returned {response_type} instead of a write-capable plan."
    return None


class GatewayHarness:
    def __init__(self, workbook: WorkbookModel, session: requests.Session):
        self.workbook = workbook
        self.session = session

    def upload_image(self, image_path: Path) -> dict[str, Any]:
        with image_path.open("rb") as handle:
            response = self.session.post(
                f"{GATEWAY_BASE_URL}/api/uploads/image",
                files={
                    "file": (image_path.name, handle, "image/png" if image_path.suffix.lower() == ".png" else "image/jpeg")
                },
                data={"source": "upload"},
                timeout=60,
            )
        response.raise_for_status()
        payload = response.json()
        return payload["attachment"]

    def build_request(self, case: DemoCase, attachments: list[dict[str, Any]] | None = None) -> dict[str, Any]:
        context = self.workbook.build_request_context(case)
        if attachments:
            context["attachments"] = attachments
        return {
            "schemaVersion": "1.0.0",
            "requestId": f"req_demo_{case.id}_{uuid.uuid4().hex[:10]}",
            "source": {
                "channel": SOURCE_CHANNEL,
                "clientVersion": CLIENT_VERSION,
                "sessionId": WORKBOOK_SESSION_KEY,
            },
            "host": {
                "platform": HOST_PLATFORM,
                "workbookTitle": self.workbook.path.name,
                "workbookId": "demo-pack-local",
                "activeSheet": case.sheet,
                "selectedRange": case.selection,
                "locale": "en-US",
                "timeZone": "Asia/Ho_Chi_Minh",
            },
            "userMessage": case.prompt,
            "conversation": [{"role": "user", "content": case.prompt}],
            "context": context,
            "capabilities": {
                "canRenderTrace": True,
                "canRenderStructuredPreview": True,
                "canConfirmWriteBack": True,
                "supportsStructureEdits": True,
                "supportsAutofit": True,
                "supportsSortFilter": True,
                "supportsImageInputs": True,
                "supportsWriteBackExecution": True,
                "supportsNoteWrites": True,
            },
            "reviewer": {
                "reviewerSafeMode": False,
                "forceExtractionMode": None,
            },
            "confirmation": {
                "state": "none",
            },
        }

    def start_run(self, request_payload: dict[str, Any]) -> dict[str, Any]:
        response = self.session.post(
            f"{GATEWAY_BASE_URL}/api/requests",
            json=request_payload,
            timeout=60,
        )
        response.raise_for_status()
        return response.json()

    def poll_run(self, run_id: str, request_id: str) -> dict[str, Any]:
        deadline = time.monotonic() + POLL_TIMEOUT_SECONDS
        last_payload: dict[str, Any] | None = None
        while time.monotonic() < deadline:
            response = self.session.get(
                f"{GATEWAY_BASE_URL}/api/requests/{run_id}",
                params={"requestId": request_id},
                timeout=30,
            )
            response.raise_for_status()
            payload = response.json()
            last_payload = payload
            if payload.get("status") in {"completed", "failed"}:
                return payload
            time.sleep(POLL_INTERVAL_SECONDS)
        raise TimeoutError(f"Timed out waiting for run {run_id}. Last payload: {last_payload}")

    def maybe_dry_run(self, response_payload: dict[str, Any], request_id: str, run_id: str) -> dict[str, Any] | None:
        response = response_payload["response"]
        if response["type"] != "composite_plan":
            return None
        dry_run_response = self.session.post(
            f"{GATEWAY_BASE_URL}/api/execution/dry-run",
            json={
                "requestId": request_id,
                "runId": run_id,
                "plan": response["data"],
                "workbookSessionKey": WORKBOOK_SESSION_KEY,
            },
            timeout=60,
        )
        dry_run_response.raise_for_status()
        return dry_run_response.json()

    def approve_and_complete(self, response_payload: dict[str, Any]) -> dict[str, Any] | None:
        response = response_payload["response"]
        response_type = response["type"]
        if response_type in {"chat", "formula", "error", "attachment_analysis", "document_summary", "extracted_table"}:
            return None

        request_id = response_payload["requestId"]
        run_id = response_payload["runId"]
        if response_type == "composite_plan":
            self.maybe_dry_run(response_payload, request_id, run_id)

        approval_payload: dict[str, Any] = {
            "requestId": request_id,
            "runId": run_id,
            "workbookSessionKey": WORKBOOK_SESSION_KEY,
            "plan": response["data"],
        }
        if response["data"].get("confirmationLevel") == "destructive":
            approval_payload["destructiveConfirmation"] = {"confirmed": True}

        approval_response = self.session.post(
            f"{GATEWAY_BASE_URL}/api/writeback/approve",
            json=approval_payload,
            timeout=60,
        )
        approval_response.raise_for_status()
        approval = approval_response.json()

        result = self.workbook.apply_plan(
            response_type=response_type,
            plan=response["data"],
            execution_id=approval.get("executionId"),
        )
        completion_response = self.session.post(
            f"{GATEWAY_BASE_URL}/api/writeback/complete",
            json={
                "requestId": request_id,
                "runId": run_id,
                "approvalToken": approval["approvalToken"],
                "planDigest": approval["planDigest"],
                "result": result,
            },
            timeout=60,
        )
        completion_response.raise_for_status()
        return {
            "approval": approval,
            "result": result,
            "completion": completion_response.json(),
        }


def load_demo_cases() -> list[DemoCase]:
    return [
        DemoCase(
            id="messy_cleanup_dashboard",
            sheet="Messy_Sales",
            selection="A1:J17",
            active_cell="A1",
            prompt="Clean this table, standardize dates/currency, remove duplicates, then create a dashboard sheet with a revenue chart.",
            expect_write=True,
            preferred_types=("composite_plan",),
        ),
        DemoCase(
            id="messy_top5_customers",
            sheet="Messy_Sales",
            selection="A1:J17",
            active_cell="A1",
            prompt="Trim customer names and standardize revenue values, then write a top 5 customers by summed revenue summary below the table. Do not remove duplicates or delete the note rows.",
            expect_write=True,
            preferred_types=("composite_plan", "sheet_update"),
        ),
        DemoCase(
            id="budget_variance_chart",
            sheet="Budget_Variance",
            selection="A1:E13",
            active_cell="A1",
            prompt="Add variance and variance percent columns, highlight overspend rows, then create a plan vs actual chart.",
            expect_write=True,
            preferred_types=("composite_plan",),
        ),
        DemoCase(
            id="marketing_roas_summary",
            sheet="Marketing_ROAS",
            selection="A1:H13",
            active_cell="A1",
            prompt="Summarize ROAS by channel, sort from best to worst, and create a revenue vs spend chart.",
            expect_write=True,
            preferred_types=("composite_plan",),
        ),
        DemoCase(
            id="inventory_restock",
            sheet="Inventory_Restock",
            selection="A1:H12",
            active_cell="A1",
            prompt="Flag low-stock items, add a reorder status column, and create a filtered urgent restock sheet.",
            expect_write=True,
            preferred_types=("composite_plan",),
        ),
        DemoCase(
            id="support_management_summary",
            sheet="Support_Tickets",
            selection="A1:H13",
            active_cell="A1",
            prompt="Create a management summary of ticket counts by priority and category, then highlight SLA risk areas.",
            expect_write=True,
            preferred_types=("composite_plan",),
        ),
        DemoCase(
            id="formula_fix_apply_h11",
            sheet="Formula_Debug",
            selection="H11",
            active_cell="H11",
            prompt="Why is the formula in H11 broken? Fix it and apply the corrected formula to H11.",
            expect_write=True,
            preferred_types=("sheet_update", "composite_plan"),
        ),
        DemoCase(
            id="formula_sumif_north",
            sheet="Formula_Debug",
            selection="H11",
            active_cell="H11",
            prompt="sumif revenue of region north",
            expect_write=False,
            preferred_types=("formula", "sheet_update"),
        ),
        DemoCase(
            id="image_sales_extract",
            sheet="Image_Import_Target",
            selection="A1",
            active_cell="A1",
            prompt="Extract this table and paste it into Image_Import_Target starting at A1.",
            image_path=DEMO_DIR / "image_sales_snapshot.png",
            expect_write=True,
            preferred_types=("sheet_import_plan", "sheet_update", "composite_plan"),
        ),
        DemoCase(
            id="image_invoice_ap_import",
            sheet="Image_Import_Target",
            selection="A1",
            active_cell="A1",
            prompt="Extract line items from this invoice image into a new sheet named AP_Import.",
            image_path=DEMO_DIR / "image_office_invoice.png",
            expect_write=True,
            preferred_types=("sheet_import_plan", "composite_plan", "workbook_structure_update"),
        ),
    ]


def ensure_environment() -> None:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(WORKBOOK_PATH, STATEFUL_WORKBOOK_PATH)


def probe_health(session: requests.Session) -> dict[str, Any]:
    gateway = session.get(f"{GATEWAY_BASE_URL}/health", timeout=15)
    gateway.raise_for_status()
    gateway_payload = gateway.json()

    brain_base = os.environ.get("HERMES_LOCAL_BRAIN_URL", "http://127.0.0.1:8642")
    brain = session.get(f"{brain_base}/health", timeout=15)
    brain.raise_for_status()
    brain_payload = brain.json()

    return {
        "gateway": gateway_payload,
        "brain": brain_payload,
    }


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")


def write_markdown_report(path: Path, summary: dict[str, Any]) -> None:
    lines: list[str] = []
    lines.append("# Hermes Demo Smoke Report")
    lines.append("")
    lines.append(f"- generatedAt: `{summary['generatedAt']}`")
    lines.append(f"- gateway: `{summary['gatewayBaseUrl']}`")
    lines.append(f"- hostPlatform: `{summary['hostPlatform']}`")
    lines.append(f"- sourceChannel: `{summary['sourceChannel']}`")
    lines.append(f"- workbookSessionKey: `{summary['workbookSessionKey']}`")
    lines.append(f"- total: `{summary['totals']['total']}`")
    lines.append(f"- passed: `{summary['totals']['passed']}`")
    lines.append(f"- failed: `{summary['totals']['failed']}`")
    lines.append("")
    lines.append("## Results")
    for case in summary["cases"]:
      status = "PASS" if case["ok"] else "FAIL"
      lines.append(f"- `{case['id']}` `{status}` `{case.get('responseType', 'unknown')}`")
      lines.append(f"  prompt: {case['prompt']}")
      if case.get("issue"):
          lines.append(f"  issue: {case['issue']}")
      if case.get("error"):
          lines.append(f"  error: {case['error']}")
      if case.get("userAction"):
          lines.append(f"  userAction: {case['userAction']}")
    path.write_text("\n".join(lines) + "\n")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run local-brain demo smoke tests for Hermes spreadsheet flows.")
    parser.add_argument(
        "--cases",
        default=os.environ.get("HERMES_DEMO_CASES", ""),
        help="Comma-separated demo case ids to run. Defaults to all."
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    ensure_environment()
    session = requests.Session()
    health = probe_health(session)
    workbook = WorkbookModel(STATEFUL_WORKBOOK_PATH)
    harness = GatewayHarness(workbook, session)
    requested_case_ids = {
        value.strip() for value in str(args.cases or "").split(",") if value.strip()
    }

    summary: dict[str, Any] = {
        "generatedAt": now_iso(),
        "gatewayBaseUrl": GATEWAY_BASE_URL,
        "hostPlatform": HOST_PLATFORM,
        "sourceChannel": SOURCE_CHANNEL,
        "workbookSessionKey": WORKBOOK_SESSION_KEY,
        "health": health,
        "cases": [],
    }

    failures = 0

    for case in load_demo_cases():
        if requested_case_ids and case.id not in requested_case_ids:
            continue
        case_dir = RESULTS_DIR / case.id
        case_dir.mkdir(parents=True, exist_ok=True)
        for stale_file in case_dir.iterdir():
            if stale_file.is_file():
                stale_file.unlink()
        print(f"[demo-smoke] {case.id}: start", flush=True)
        attachments = None
        try:
            if case.image_path:
                attachments = [harness.upload_image(case.image_path)]

            request_payload = harness.build_request(case, attachments=attachments)
            write_json(case_dir / "request.json", request_payload)

            start_payload = harness.start_run(request_payload)
            write_json(case_dir / "accepted.json", start_payload)

            final_payload = harness.poll_run(start_payload["runId"], start_payload["requestId"])
            write_json(case_dir / "final.json", final_payload)

            response = final_payload.get("response")
            error_payload = final_payload.get("error")
            issue = None
            approval_completion = None
            ok = False
            response_type = None
            user_action = None

            if final_payload.get("status") == "failed" or error_payload:
                message = error_payload.get("message") if isinstance(error_payload, dict) else "Run failed."
                user_action = error_payload.get("userAction") if isinstance(error_payload, dict) else None
                issue = f"Run failed before a valid response: {message}"
            elif not response:
                issue = "Run completed without a response payload."
            else:
                response_type = response["type"]
                issue = expectation_issue(case, response_type)
                if response_type == "error":
                    issue = response["data"]["message"]
                    user_action = response["data"].get("userAction")
                else:
                    if response_type not in {"chat", "formula", "attachment_analysis", "document_summary", "extracted_table"}:
                        approval_completion = harness.approve_and_complete(final_payload)
                        if approval_completion is not None:
                            write_json(case_dir / "writeback.json", approval_completion)
                            workbook.save()
                    ok = issue is None

            if not ok:
                failures += 1

            summary["cases"].append({
                "id": case.id,
                "prompt": case.prompt,
                "sheet": case.sheet,
                "ok": ok,
                "responseType": response_type,
                "issue": issue,
                "userAction": user_action,
                "runId": start_payload["runId"],
                "requestId": start_payload["requestId"],
                "finalStatus": final_payload.get("status"),
                "requiresConfirmation": response.get("ui", {}).get("showRequiresConfirmation") if response else None,
                "writebackCompleted": approval_completion is not None,
            })

        except Exception as exc:  # pragma: no cover - harness safety
            failures += 1
            error_payload: dict[str, Any] = {
                "error": str(exc),
                "type": exc.__class__.__name__,
            }
            response = getattr(exc, "response", None)
            if response is not None:
                try:
                    error_payload["statusCode"] = response.status_code
                    error_payload["body"] = response.json()
                except Exception:
                    error_payload["statusCode"] = getattr(response, "status_code", None)
                    error_payload["body"] = getattr(response, "text", "")
                write_json(case_dir / "http_error.json", error_payload)
            summary["cases"].append({
                "id": case.id,
                "prompt": case.prompt,
                "sheet": case.sheet,
                "ok": False,
                "error": str(exc),
            })
        else:
            print(f"[demo-smoke] {case.id}: done", flush=True)

    summary["totals"] = {
        "total": len(summary["cases"]),
        "passed": len(summary["cases"]) - failures,
        "failed": failures,
    }

    write_json(RESULTS_DIR / "summary.json", summary)
    write_markdown_report(RESULTS_DIR / "REPORT.md", summary)
    workbook.save()

    print(json.dumps(summary["totals"], indent=2))
    print(f"Detailed results: {RESULTS_DIR}")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main())
